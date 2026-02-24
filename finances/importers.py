import csv
import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError

from .models import Category, Transaction


TYPE_ALIASES = {
    'income': Transaction.Type.INCOME,
    'receita': Transaction.Type.INCOME,
    'in': Transaction.Type.INCOME,
    'expense': Transaction.Type.EXPENSE,
    'despesa': Transaction.Type.EXPENSE,
    'out': Transaction.Type.EXPENSE,
}


def normalize_token(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return text


def normalize_header(value):
    return normalize_token(value).replace(' ', '_')


def read_uploaded_rows(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith('.csv'):
        return _read_csv_rows(uploaded_file)
    if filename.endswith('.xlsx'):
        return _read_xlsx_rows(uploaded_file)
    raise ValueError('Formato de arquivo invalido. Use CSV ou XLSX.')


def _read_csv_rows(uploaded_file):
    raw = uploaded_file.read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return []

    normalized_headers = [normalize_header(field_name) for field_name in reader.fieldnames]
    rows = []
    for row_index, row_data in enumerate(reader, start=2):
        mapped_row = {}
        for original_name, normalized_name in zip(reader.fieldnames, normalized_headers):
            mapped_row[normalized_name] = row_data.get(original_name)
        rows.append((row_index, mapped_row))

    return rows


def _read_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Biblioteca openpyxl nao encontrada para leitura de XLSX.') from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    iter_rows = sheet.iter_rows(values_only=True)

    try:
        headers = next(iter_rows)
    except StopIteration:
        return []

    normalized_headers = [normalize_header(header) for header in headers]
    rows = []
    for row_index, values in enumerate(iter_rows, start=2):
        row_map = {}
        for index, header in enumerate(normalized_headers):
            value = values[index] if index < len(values) else None
            row_map[header] = value
        rows.append((row_index, row_map))

    workbook.close()
    return rows


def _parse_type(raw_value):
    normalized = normalize_token(raw_value)
    if normalized in TYPE_ALIASES:
        return TYPE_ALIASES[normalized]
    normalized = normalized.upper()
    if normalized in {Transaction.Type.INCOME, Transaction.Type.EXPENSE}:
        return normalized
    return None


def _parse_amount(raw_value):
    if raw_value is None:
        raise ValueError('valor ausente')

    if isinstance(raw_value, Decimal):
        amount = raw_value
    elif isinstance(raw_value, (int, float)):
        amount = Decimal(str(raw_value))
    else:
        text = str(raw_value).strip().replace('R$', '').replace(' ', '')
        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        elif ',' in text:
            text = text.replace('.', '').replace(',', '.')

        try:
            amount = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError('valor invalido') from exc

    if amount <= 0:
        raise ValueError('valor deve ser maior que zero')

    return amount


def _parse_date(raw_value):
    if raw_value is None:
        raise ValueError('data ausente')

    if isinstance(raw_value, datetime):
        return raw_value.date()

    if isinstance(raw_value, date):
        return raw_value

    text = str(raw_value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError('data invalida, use YYYY-MM-DD ou DD/MM/YYYY')


def _parse_bool(raw_value):
    if raw_value in {None, ''}:
        return False
    if isinstance(raw_value, bool):
        return raw_value

    normalized = normalize_token(raw_value)
    return normalized in {'1', 'true', 'sim', 'yes', 'y', 's'}


def _get_value(row, *keys):
    for key in keys:
        if key in row and row[key] not in {None, ''}:
            return row[key]
    return None


def import_categories_from_file(uploaded_file, user):
    rows = read_uploaded_rows(uploaded_file)
    created_count = 0
    skipped_count = 0
    errors = []

    for row_number, row in rows:
        name = _get_value(row, 'name', 'nome')
        type_value = _get_value(row, 'type', 'tipo')

        if not name or not type_value:
            errors.append(f'Linha {row_number}: campos obrigatorios name e type.')
            continue

        parsed_type = _parse_type(type_value)
        if not parsed_type:
            errors.append(f'Linha {row_number}: tipo invalido ({type_value}).')
            continue

        category_name = str(name).strip()
        if not category_name:
            errors.append(f'Linha {row_number}: nome da categoria vazio.')
            continue

        if Category.objects.filter(user=user, type=parsed_type, name__iexact=category_name).exists():
            skipped_count += 1
            continue

        try:
            Category.objects.create(user=user, name=category_name, type=parsed_type)
            created_count += 1
        except ValidationError as exc:
            errors.append(f'Linha {row_number}: {exc.messages[0]}')

    return {
        'total_rows': len(rows),
        'created': created_count,
        'skipped': skipped_count,
        'errors': errors,
    }


def import_transactions_from_file(uploaded_file, user):
    rows = read_uploaded_rows(uploaded_file)
    created_count = 0
    errors = []

    categories_by_name_and_type = {}
    categories_by_id_and_type = {}
    for category in Category.objects.filter(user=user):
        key = (normalize_token(category.name), category.type)
        categories_by_name_and_type[key] = category
        categories_by_id_and_type[(str(category.id), category.type)] = category

    for row_number, row in rows:
        raw_type = _get_value(row, 'type', 'tipo')
        raw_amount = _get_value(row, 'amount', 'valor')
        raw_date = _get_value(row, 'date', 'data')
        raw_category = _get_value(row, 'category', 'categoria')
        raw_description = _get_value(row, 'description', 'descricao') or ''
        raw_notes = _get_value(row, 'notes', 'observacao') or ''
        raw_recurring = _get_value(row, 'is_recurring', 'recorrente')

        if not raw_type or raw_amount in {None, ''} or not raw_date or not raw_category:
            errors.append(
                f'Linha {row_number}: campos obrigatorios type, amount, date e category.'
            )
            continue

        parsed_type = _parse_type(raw_type)
        if not parsed_type:
            errors.append(f'Linha {row_number}: tipo invalido ({raw_type}).')
            continue

        try:
            parsed_amount = _parse_amount(raw_amount)
        except ValueError as exc:
            errors.append(f'Linha {row_number}: {exc}.')
            continue

        try:
            parsed_date = _parse_date(raw_date)
        except ValueError as exc:
            errors.append(f'Linha {row_number}: {exc}.')
            continue

        category_key_by_id = (str(raw_category).strip(), parsed_type)
        category_key_by_name = (normalize_token(raw_category), parsed_type)
        category = categories_by_id_and_type.get(category_key_by_id)
        if not category:
            category = categories_by_name_and_type.get(category_key_by_name)

        if not category:
            errors.append(
                f'Linha {row_number}: categoria "{raw_category}" nao encontrada para o tipo selecionado.'
            )
            continue

        try:
            Transaction.objects.create(
                user=user,
                type=parsed_type,
                amount=parsed_amount,
                date=parsed_date,
                category=category,
                description=str(raw_description).strip(),
                notes=str(raw_notes).strip(),
                is_recurring=_parse_bool(raw_recurring),
            )
            created_count += 1
        except ValidationError as exc:
            errors.append(f'Linha {row_number}: {exc.messages[0]}')

    return {
        'total_rows': len(rows),
        'created': created_count,
        'errors': errors,
    }
